import xlrd
import openpyxl
import sys
import os
import xlwt

def calculate_vm_count(build_plan_path):
	"""
	This function:
		- gathers the "# of VM's" per VM type for file generation 
		- creates list of properly named titles for final output folder based off of each vm-types vf-module-name
		- used for function calls at the end of the program
	"""
	global final_vf_module_name
	global title_list

	# gather the vm_type
	pt = openpyxl.load_workbook(preload_path)
	ws = pt.get_sheet_by_name(u'VMs')
	vm_type = ws['B7'].value
	#print("VM TYPE: ", vm_type)    

	# set variable to hold vm count
	global vm_count 
	# open workbook and specify which sheet you would like to access
	wb = xlrd.open_workbook(build_plan_path)
	sheet_names = wb.sheet_names()
	bp = wb.sheet_by_name(u'VNF-Specs')

	# string search VNF-Specs column headers and assign each columns reference position (int) to a variable
	# this avoids hard coding the position of certain columns
	# order does not matter
	for i in range(bp.nrows):
		for j in range(bp.ncols):
			if bp.cell_value(i, j) == "vm-type":
				col_ref_vmt = j
			if bp.cell_value(i, j) == "# of VM's":
				col_ref_vmc = j
			if bp.cell_value(i, j) == "vf-module-name":
					col_ref_vfmn = j

	# create dict of vm-type and each vm-types VM Count
	count_dict = {}
	for i in range(bp.nrows):
		vm = bp.cell_value(i, col_ref_vmt)
		count = bp.cell_value(i, col_ref_vmc)
		count_dict[vm] = count

	# cast as integer as it is pulled as a string
	for k, v in count_dict.items():
		#print("TYPE:", vm_type)
		if k == vm_type:
			vm_count = int(v)

	# create dictionary of vm-names and vfmn for title generation  
	vfmn_dict = {}
	for i in range(bp.nrows):
		vm = bp.cell_value(i, col_ref_vmt)
		vfmn = bp.cell_value(i, col_ref_vfmn)
		vfmn_dict[vm] = vfmn

	# give the titles based on which vm-type the current file is
	title_list = []
	for k, v in vfmn_dict.items():
		for i in range(1, vm_count+1):
			if k != '' and k == vm_type:
				title_list.append(v)
	#print(title_list)


def change_general(preload_path, build_plan_path, count):
	"""
	This function:
		- initiates changes for General tab in preload template
	"""
	# find module type
	pt = openpyxl.load_workbook(preload_path)
	ws = pt.get_sheet_by_name(u'VMs')
	vm_type = ws['B7'].value
	print("VM TYPE:", vm_type)

	# open workbook and specify which sheet you would like to access
	wb = xlrd.open_workbook(build_plan_path)
	sheet_names = wb.sheet_names()
	bp = wb.sheet_by_name(u'VNF-Specs')

	# string search VNF-Specs column headers and assign each columns reference position (int) to a variable
	# this avoids hard coding the position of certain columns
	# order does not matter
	for i in range(bp.nrows):
		for j in range(bp.ncols):
			if bp.cell_value(i, j) == "vm-type":
				col_ref_vmt = j
			if bp.cell_value(i, j) == "# of VM's":
				col_ref_vmc = j
			if bp.cell_value(i, j) == "vf-module-name":
				col_ref_vfmn = j
			if bp.cell_value(i, j) == "vf-module-model-name":
				col_ref_vfmmn = j
			if bp.cell_value(i, j) == "vnf-name":
				col_ref_vnfn = j
			if bp.cell_value(i, j) == "vnf-type":
				col_ref_vnft = j
			if bp.cell_value(i, j) == "vf-module-model-name-base":
				col_ref_vfmmnb = j
			if bp.cell_value(i, j) == "probe_pod":
				col_ref_probe_prod = j
	#print("vmt: ", col_ref_vmt)
	#print("vmc: ", col_ref_vmc)

	# creates dict of vm-types and vf-module-names
	vf_module_name_dict = {}
	for i in range(bp.nrows):
		vm = bp.cell_value(i, col_ref_vmt)
		modules = bp.cell_value(i, col_ref_vfmn)
		vf_module_name_dict[vm] = modules
	#print(vf_module_name_dict)

	# creates dict of vm-types and vf-module-model-names
	vf_module_model_name_dict = {}
	for i in range(bp.nrows):
		vm = bp.cell_value(i, col_ref_vmt)
		modules_model = bp.cell_value(i, col_ref_vfmmn)
		vf_module_model_name_dict[vm] = modules_model

	# creates dict of vm-types and vf-module-model-name-base
	vf_module_model_name_base_dict = {}
	for i in range(bp.nrows):
		vm = bp.cell_value(i, col_ref_vmt)
		modules_model_base = bp.cell_value(i, col_ref_vfmmnb)
		vf_module_model_name_base_dict[vm] = modules_model_base

	# creates dict of vm-types and vnf-names
	vnf_name_dict = {}
	for i in range(bp.nrows):
		vm = bp.cell_value(i, col_ref_vmt)
		vnf_name = bp.cell_value(i, col_ref_vnfn)
		vnf_name_dict[vm] = vnf_name

	# creates dict of vm-type and vnf-types
	vnf_type_dict = {}
	for i in range(bp.nrows):
		vm = bp.cell_value(i, col_ref_vmt)
		vnf_type = bp.cell_value(i, col_ref_vnft)
		vnf_type_dict[vm] = vnf_type

	# update vf-module-name
	wb = openpyxl.load_workbook(preload_path)
	ws = wb.get_sheet_by_name(u'General')

	for k, v in vf_module_name_dict.items():
		if k == vm_type:
			#print(v + count)
			ws['C6'].value = v + count
			#wb.sheets[1].range('C6').value = v + count # proper name # SUFFIX

	#update vf module model
	#wb = openpyxl.load_workbook(preload_path)
	#ws = wb.get_sheet_by_name(u'General')
	for k, v in vf_module_model_name_dict.items():
		if k == vm_type:
			#print(v)
			ws['C8'].value = v

	# update vnf-name
	#wb = openpyxl.load_workbook(preload_path)
	#ws = wb.get_sheet_by_name(u'General')
	for k, v in vnf_name_dict.items():
		if k == vm_type:
			#print(v)
			ws['C12'].value = v

	# update vnf-type
	#wb = openpyxl.load_workbook(preload_path)
	#ws = wb.get_sheet_by_name(u'General')
	ws['C14'] = ""
	for k, v in vnf_type_dict.items():
		if k == vm_type:
			#print(v)
			ws['C13'].value = v
	wb.save(dest_folder + title_list[titles] + str(titles+1) + ".xlsm")
##########################################################################################################################



build_plan_path = sys.argv[1]
preload_list = []
paths = sys.argv[2]
for idx, item in enumerate(os.listdir(paths)):
	preload_list.append(paths+item)
#print(preload_list)
dest_folder = sys.argv[3]

files = []
for preload_path in preload_list:
	if "base" not in preload_path:
		calculate_vm_count(build_plan_path)
		for titles in range(0, len(title_list)):
			wb = openpyxl.load_workbook(preload_path)
			change_general(preload_path, build_plan_path, str(titles + 1))
			#wb.save(dest_folder + title_list[titles] + str(titles+1) + ".xlsm")
			#print(dest_folder + title_list[titles] + str(titles+1) + ".xlsm")
