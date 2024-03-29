import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import sys
from shutil import make_archive
import os
import datetime
import re
import warnings

warnings.filterwarnings('ignore')  # supress UserWarning Excel Data Validation

# 8-24-21

def calculate_vm_count(build_plan_path):
    """
	This function:
		- gathers the "# of VM's" per VM type for file generation 
		- creates list of properly named titles for final output folder based off of each vm-types vf-module-name
		- used for function calls at the end of the program
	"""
    global final_vf_module_name
    global title_list

    # gather the vm_type
    pt = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = pt[u'VMs']
    vm_type = ws['B7'].value

    # set variable to hold vm count
    global vm_count
    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'VNF-Specs')

    # string search VNF-Specs column headers and assign each columns reference position (int) to a variable
    # this avoids hard coding the position of certain columns
    # order does not matter
    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "vm-type":
                col_ref_vmt = j
            if bp.cell_value(i, j) == "# of VM's":
                col_ref_vmc = j
            if bp.cell_value(i, j) == "vf-module-name":
                col_ref_vfmn = j
    # print("vmc: ", col_ref_vmc)

    # create dict of vm-type and each vm-types VM Count
    count_dict = {}
    for i in range(bp.nrows):
        vm = bp.cell_value(i, col_ref_vmt)
        vm_count = bp.cell_value(i, col_ref_vmc)
        count_dict[vm] = vm_count

    # print("cd: ", count_dict)

    # cast as integer as it is pulled as a string
    for k, v in count_dict.items():
        if k != '' and k == vm_type:
            vm_count = int(float(v))

    # print("vmcount: ", type(vm_count))

    # create dictionary of vm-names and vfmn for title generation
    vfmn_dict = {}
    for i in range(bp.nrows):
        vm = bp.cell_value(i, col_ref_vmt)
        vfmn = bp.cell_value(i, col_ref_vfmn)
        vfmn_dict[vm] = vfmn

    # print("vfmn_dict: ", vfmn_dict)

    # give the titles based on which vm-type the current file is
    title_list = []
    for k, v in vfmn_dict.items():
        for i in range(1, int(vm_count) + 1):
            if k != '' and k == vm_type:
                title_list.append(v)


####### CHANGE GENERAL ########
def change_general(preload_path, build_plan_path, count):
    """
	This function:
		- initiates changes for General tab in preload template
	"""
    # find module type
    pt = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = pt[u'VMs']
    vm_type = ws['B7'].value

    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'VNF-Specs')

    # string search VNF-Specs column headers and assign each columns reference position (int) to a variable
    # this avoids hard coding the position of certain columns
    # order does not matter
    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "vm-type":
                col_ref_vmt = j
            if bp.cell_value(i, j) == "# of VM's":
                col_ref_vmc = j
            if bp.cell_value(i, j) == "vf-module-name":
                col_ref_vfmn = j
            if bp.cell_value(i, j) == "vf-module-model-name":
                col_ref_vfmmn = j
            if bp.cell_value(i, j) == "vnf-name":
                col_ref_vnfn = j
            if bp.cell_value(i, j) == "vnf-type":
                col_ref_vnft = j
            if bp.cell_value(i, j) == "vf-module-model-name-base":
                col_ref_vfmmnb = j
            if bp.cell_value(i, j) == "probe_pod":
                col_ref_probe_prod = j

    # creates dict of vm-types and vf-module-names
    vf_module_name_dict = {}
    for i in range(bp.nrows):
        vm = bp.cell_value(i, col_ref_vmt)
        modules = bp.cell_value(i, col_ref_vfmn)
        vf_module_name_dict[vm] = modules

    # creates dict of vm-types and vf-module-model-names
    vf_module_model_name_dict = {}
    for i in range(bp.nrows):
        vm = bp.cell_value(i, col_ref_vmt)
        modules_model = bp.cell_value(i, col_ref_vfmmn)
        vf_module_model_name_dict[vm] = modules_model

    # creates dict of vm-types and vf-module-model-name-base
    vf_module_model_name_base_dict = {}
    for i in range(bp.nrows):
        vm = bp.cell_value(i, col_ref_vmt)
        modules_model_base = bp.cell_value(i, col_ref_vfmmnb)
        vf_module_model_name_base_dict[vm] = modules_model_base

    # creates dict of vm-types and vnf-names
    vnf_name_dict = {}
    for i in range(bp.nrows):
        vm = bp.cell_value(i, col_ref_vmt)
        vnf_name = bp.cell_value(i, col_ref_vnfn)
        vnf_name_dict[vm] = vnf_name

    # creates dict of vm-type and vnf-types
    vnf_type_dict = {}
    for i in range(bp.nrows):
        vm = bp.cell_value(i, col_ref_vmt)
        vnf_type = bp.cell_value(i, col_ref_vnft)
        vnf_type_dict[vm] = vnf_type

    wb = openpyxl.load_workbook(preload_path, keep_vba=True, read_only=False)
    ws = wb[u'General']

    # update vf-module-name
    for k, v in vf_module_name_dict.items():
        if k == vm_type:
            ws['C6'].value = v + count

    # update vf module model
    for k, v in vf_module_model_name_dict.items():
        if k == vm_type:
            ws['C8'].value = v

    # update vnf-name
    for k, v in vnf_name_dict.items():
        if k == vm_type:
            ws['C12'].value = v

    # update vnf-type
    for k, v in vnf_type_dict.items():
        if k == vm_type:
            ws['C13'].value = v

    wb.save(preload_path)


def change_networks(preload_path, build_plan_path):
    """
	This function:
		- initiates changes for Networks information
	"""
    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'Networks')

    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "network role":
                col_ref_nr = j
            if bp.cell_value(i, j) == "Network Name":
                col_ref_nn = j
            if bp.cell_value(i, j) == "Subnet_Name":
                col_ref_sn = j

    # creates dict of network_role and network_name
    # these column references are hard coded
    net_dict = {}
    for i in range(bp.nrows):
        network_role = bp.cell_value(i, col_ref_nr)
        network_name = bp.cell_value(i, col_ref_nn)
        net_dict[network_role] = network_name

    # create dict for network role and subnet_name
    subnet_dict = {}
    for i in range(bp.nrows):
        network_role = bp.cell_value(i, col_ref_nr)
        subnet_name = bp.cell_value(i, col_ref_sn)
        subnet_dict[network_role] = subnet_name

    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(preload_path)
    sheet_names = wb.sheet_names()
    networks_sheet = wb.sheet_by_name(u'Networks')

    # implement changes to template
    wb = openpyxl.load_workbook(preload_path, keep_vba=True, read_only=False)
    ws = wb[u'Networks']
    for k, v in net_dict.items():
        for i in range(networks_sheet.nrows):
            if (networks_sheet.cell_value(i, 1) == k and k != ''):
                val = str(i + 1)
                ws['C' + val].value = v

    # implement changes to template
    for k, v in subnet_dict.items():
        for i in range(networks_sheet.nrows):
            if (networks_sheet.cell_value(i, 1) == k and k != ''):
                val = str(i + 1)
                ws['F' + val].value = v

    wb.save(preload_path)


def change_probe_prod(preload_path, build_plan_path):
    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'VNF-Specs')

    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "vm-type":
                col_ref_vmt = j
            if bp.cell_value(i, j) == "probe_pod":
                col_ref_probe_prod = j
            else:
                col_ref_probe_prod = 0

    if col_ref_probe_prod > 0 :
        # creates dict of vm-types and vf-module-names
        probe_pod_dict = {}
        for i in range(bp.nrows):
            vm = bp.cell_value(i, col_ref_vmt)
            probe_pod = bp.cell_value(i, col_ref_probe_prod)
            probe_pod_dict[vm] = probe_pod

        # update probe pod in tag values NOT general tab
        # get vm type
        # scan through common parameters, replace cell next to probe_prod
        # open workbook and specify which sheet you would like to access

        pt = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
        ws = pt[u'VMs']
        vm_type_for_probe_prod = ws['B7'].value

        wb = xlrd.open_workbook(preload_path)
        sheet_names = wb.sheet_names()
        tag_sheet = wb.sheet_by_name(u'Tag-values')

        wb = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
        ws = wb[u'Tag-values']
        for i in range(tag_sheet.nrows):
            if (tag_sheet.cell_value(i, 1) == "probe_pod"):
                for k, v in probe_pod_dict.items():
                    if k == vm_type_for_probe_prod:
                        val = str(i + 1)
                        ws['C' + val].value = v
        wb.save(preload_path)


def change_vccn(preload_path, build_plan_path):
    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'VNF-Specs')

    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "vm-type":
                col_ref_vmt = j
            if bp.cell_value(i, j) == "vertica_configuration_cluster_name":
                col_ref_vccn = j

    # creates dict of vm-types and vf-module-names
    vccn_dict = {}
    for i in range(bp.nrows):
        try:
            vm = bp.cell_value(i, col_ref_vmt)
            vccn = bp.cell_value(i, col_ref_vccn)
            vccn_dict[vm] = vccn
        except:
            pass

    # update probe pod in tag values NOT general tab
    # get vm type
    # scan through common parameters, replace cell next to probe_prod
    # open workbook and specify which sheet you would like to access
    pt = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = pt[u'VMs']
    vm_type_for_vccn = ws['B7'].value

    wb = xlrd.open_workbook(preload_path)
    sheet_names = wb.sheet_names()
    tag_sheet = wb.sheet_by_name(u'Tag-values')

    wb = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = wb[u'Tag-values']
    for i in range(tag_sheet.nrows):
        if (tag_sheet.cell_value(i, 1) == "vertica_configuration_cluster_name"):
            for k, v in vccn_dict.items():
                if k == vm_type_for_vccn:
                    val = str(i + 1)
                    ws['C' + val].value = v
    wb.save(preload_path)


def change_tag(preload_path, build_plan_path):
    """
	This function:
		- initiates changes for all tag values
	"""
    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'Common Parameters')

    params = []
    for i in range(0, bp.nrows):
        if bp.cell_value(i, 0) == "Parameter Name":
            params.append(i)
    start = int(params[-1]) + 1

    # create dict of parameter name and associated value
    tag_dict = {}
    for i in range(start, bp.nrows):
        par = bp.cell_value(i, 0)
        val = bp.cell_value(i, 1)
        tag_dict[par] = val

    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(preload_path)
    sheet_names = wb.sheet_names()
    tag_sheet = wb.sheet_by_name(u'Tag-values')

    # implement changes to template
    wb = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = wb[u'Tag-values']
    for k, v in tag_dict.items():
        for i in range(tag_sheet.nrows):
            if (tag_sheet.cell_value(i, 1) == k and k != ''):
                val = str(i + 1)
                ws['C' + val].value = v
    wb.save(preload_path)


def change_vm(preload_path, build_plan_path, count):
    """
	This function:
		- initates changes for VM's tab
	"""
    # check vm-type
    pt = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = pt[u'VMs']
    vm_type = ws['B7'].value

    # grab values for vm-name and calculate appropriate suffix
    bp = openpyxl.load_workbook(build_plan_path, read_only=False, keep_vba=True)
    ws = bp[u'Site-Info']
    is_different = ws['B15'].value
    # print("is different: ", is_different)
    if is_different == None or is_different != True:
        wb = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
        ws = wb[u'General']
        vnf_name_general = ws['C12'].value
    else:
        bp = openpyxl.load_workbook(build_plan_path, read_only=False, keep_vba=True)
        ws = bp[u'Site-Info']
        vnf_name_general = ws['C15'].value
    # print(vnf_name_general)

    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'VM-Layout')

    # string search VM-Layout column headers and assign each columns reference position (int) to a variable
    # this avoids hard coding the position of certain columns
    # order does not matter
    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "vm-type":
                col_ref_vmt = j
            if bp.cell_value(i, j) == "VFC ID (ppp)":
                col_ref_ppp = j

    # creates dict of vm_names and ppp
    ppp_dict = {}
    for i in range(bp.nrows):
        vm_types = bp.cell_value(i, col_ref_vmt)
        ppp = bp.cell_value(i, col_ref_ppp)
        ppp_dict[vm_types] = ppp

    # instantiate replacements
    for k, v in ppp_dict.items():
        if vm_type == k:
            if int(count) < 10:
                vm_name_val = vnf_name_general + v + "00" + count
            else:
                vm_name_val = vnf_name_general + v + "0" + count
    # print("VAL: ", preload_path, vm_name_val)

    wb = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = wb[u'VMs']
    ws['C7'].value = vm_name_val

    wb.save(preload_path)


def change_az(preload_path, build_plan_path):
    """
	This function:
		- initiates changes for AZ's
	"""
    # open workbook and specify which sheet you would  to access
    # save vm_name
    pt = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = pt[u'VMs']
    vm_name_value = ws['C7'].value

    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'VM-Layout')

    # string search VM-Layout column headers and assign each columns reference position (int) to a variable
    # this avoids hard coding the position of certain columns
    # order does not matter
    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "vm-name":
                col_ref_vmn = j
            if bp.cell_value(i, j) == "AZ:Compute":
                col_ref_azc = j

    # creates dict of vm_names and az's
    az_dict = {}
    for i in range(bp.nrows):
        vm_names = bp.cell_value(i, col_ref_vmn)
        az = bp.cell_value(i, col_ref_azc)
        az_dict[vm_names] = az

    # instantiates changes based on key, replaces cell with value
    wb = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = wb[u'Availability-zones']
    for k, v in az_dict.items():
        if k == vm_name_value:
            ws['B6'].value = v

    wb.save(preload_path)


def change_vm_network_ips(preload_path, build_plan_path):
    """
	This function:
			- fills proper information into VM-network-IPs
	"""
    # extract vm-name
    pt = openpyxl.load_workbook(preload_path)
    ws = pt[u'VMs']
    vm_name = ws['C7'].value

    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'VM-Layout')

    # string search VM-Layout column headers and assign each columns reference position (int) to a variable
    # this avoids hard coding the position of certain columns
    # order does not matter
    col_ref_ngxp = {}
    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "vm-name":
                col_ref_vmn = j
            if bp.cell_value(i, j) == "oam_protected":
                col_ref_oam = j
            else:  # ngxp
                ngxp_net_cols = ["ngxp_oam_net", "ngxp_mgt_net", "ngxp_cdr_net", "int_ngxp_be_ccd_net",
                                 "int_ngxp_admin_ccd_net", "int_ngxp_ecde_ccd_net", "oam_protected_net",
                                 "int_ngxp_fe_ccd_net", "int_ngxp_pcu_net", "int_probe_agent_net_1", "int_probe_agent_net_2"]
                for val in ngxp_net_cols:
                    if bp.cell_value(i, j) == val:
                        col_ref_ngxp[bp.cell_value(i, j)] = j
    # create dictionary
    oam_dict = {}
    if col_ref_ngxp:  # ngxp
        for i in range(bp.nrows):
            oam_ip_dict = {}
            vmn = bp.cell_value(i, col_ref_vmn)
            for key, idx in col_ref_ngxp.items():
                try:
                    oam = bp.cell_value(i, idx)
                    oam_ip_dict[key].append(oam)
                    oam_dict[vmn].append(oam_ip_dict)  # store IPs as dict of dictionary
                except KeyError:
                    oam_ip_dict[key] = [oam]
                    oam_dict[vmn] = [oam_ip_dict]
    else:
        for i in range(bp.nrows):
            vmn = bp.cell_value(i, col_ref_vmn)
            oam = bp.cell_value(i, col_ref_oam)
            oam_dict[vmn] = oam

    # replace values
    wb = openpyxl.load_workbook(preload_path)
    ws = wb[u'VM-network-IPs']
    if col_ref_ngxp:  # ngxp
        col_ref_ntw = [i for i, col in enumerate(ws.iter_cols()) for cell in col if cell.value == 'network-role'][0]
        ip_dict = [v for k, v in oam_dict.items() if k == vm_name][0]  # ip dict of dictionary
        for idx, row in enumerate(ws.rows):
            if row[col_ref_ntw].value and row[col_ref_ntw].value != 'network-role':
                ip_to_match = str(row[col_ref_ntw].value)[:-1]
                # match IP from network-role to VM-Layout
                ip_match = [val[0] for d in ip_dict for key, val in d.items() if re.findall(ip_to_match, key)][0]
                ws.cell(idx + 1, col_ref_ntw + 2).value = ip_match
    else:
        for k, v in oam_dict.items():
            if k == vm_name:
                ws['D7'].value = v
        # wb = xw.Book(preload_path)
        # wb.sheets[6].range('D7').value = v
    wb.save(preload_path)


def names_tag_sheet_nf_naming(preload_path, build_plan_path):
    """
	This function:
		- creates a list of comma seperated names based on vvvv and ppp

	"""
    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'VM-Layout')

    # string search VM-Layout column headers and assign each columns reference position (int) to a variable
    # this avoids hard coding the position of certain columns
    # order does not matter
    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "vm-name":
                col_ref_vmn = j
            if bp.cell_value(i, j) == "nf_naming_code (VF Level) VVVV Code":
                col_ref_vvvv = j
            if bp.cell_value(i, j) == "vm-type":
                col_ref_vmt = j

    rlba_list = []
    rvlb_list = []
    rprb_list = []
    rqrt_list = []

    for i in range(bp.nrows):
        if "rlba" in bp.cell_value(i, col_ref_vmn):
            rlba_list.append(bp.cell_value(i, col_ref_vmn))
        if "rvlb" in bp.cell_value(i, col_ref_vmn):
            rvlb_list.append(bp.cell_value(i, col_ref_vmn))
        if "rprb" in bp.cell_value(i, col_ref_vmn):
            rprb_list.append(bp.cell_value(i, col_ref_vmn))
        if "rqrt" in bp.cell_value(i, col_ref_vmn):
            rqrt_list.append(bp.cell_value(i, col_ref_vmn))

    rlba_list = ('[%s]' % ','.join(map(str, rlba_list)))[1:-1]
    rvlb_list = ('[%s]' % ','.join(map(str, rvlb_list)))[1:-1]
    rprb_list = ('[%s]' % ','.join(map(str, rprb_list)))[1:-1]
    rqrt_list = ('[%s]' % ','.join(map(str, rqrt_list)))[1:-1]

    vvvv_dict = {
        "vlbagent": rlba_list,
        "vlb": rvlb_list,
        "vprb": rprb_list,
        "qrouter": rqrt_list
    }

    # take vm type
    pt = openpyxl.load_workbook(preload_path)
    ws = pt[u'VMs']
    vm_type = ws['B7'].value
    vm_type = str(vm_type)

    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(preload_path)
    sheet_names = wb.sheet_names()
    tag_sheet = wb.sheet_by_name(u'Tag-values')

    wb = openpyxl.load_workbook(preload_path)
    ws = wb[u'Tag-values']
    for i in range(tag_sheet.nrows):
        if (tag_sheet.cell_value(i, 1) == (str(vm_type) + "_names")):
            for k, v in vvvv_dict.items():
                if k == vm_type:
                    val = str(i + 1)
                    # print("str(v)! :", v)
                    ws['C' + val].value = str(v)  # should this be string
    wb.save(preload_path)


def names_tag_sheet(preload_path, build_plan_path):
    """
	This function:
		- creates the list of comma seperated names for Tag-values sheet
	"""
    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'VM-Layout')

    # string search VM-Layout column headers and assign each columns reference position (int) to a variable
    # this avoids hard coding the position of certain columns
    # order does not matter
    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "vm-name":
                col_ref_vmn = j
            if bp.cell_value(i, j) == "VFC ID (ppp)":
                col_ref_ppp = j
            if bp.cell_value(i, j) == "vm-type":
                col_ref_vmt = j

    # creates lists of all the names
    prb_list = []
    qrt_list = []
    lba_list = []
    vlbeph_list = []
    vlb_list = []
    ana_list = []
    msr_list = []
    cdp_list = []
    aku_list = []
    mbm_list = []
    ttn_list = []
    mgu_list = []
    con_list = []
    qtp_list = []
    ccm_list = []
    qlb_list = []
    gdn_list = []
    dbm_list = []
    akf_list = []
    dtl_list = []
    ssr_list = []
    vdb_list = []
    log_list = []
    imm_list = []
    srp_list = []
    crp_list = []
    shd_list = []
    ldr_list = []
    cgw_list = []
    dmn_list = []
    agw_list = []
    ricn_list = []
    cpt_list = []
    ssf_list = []
    k8m_list = []
    k8w_list = []
    cke_list = []
    keb_list = []
    fmc_list = []
    isb_list = []

    for i in range(bp.nrows):
        if "ricn" in bp.cell_value(i, col_ref_vmn):
            ricn_list.append(bp.cell_value(i, col_ref_vmn))
        if "prb" in bp.cell_value(i, col_ref_vmn):
            prb_list.append(bp.cell_value(i, col_ref_vmn))
        if "qrt" in bp.cell_value(i, col_ref_vmn):
            qrt_list.append(bp.cell_value(i, col_ref_vmn))
        if "lba" in bp.cell_value(i, col_ref_vmn):
            lba_list.append(bp.cell_value(i, col_ref_vmn))
        if "vlb" in bp.cell_value(i, col_ref_vmn):
            vlb_list.append(bp.cell_value(i, col_ref_vmn))
        if "ana" in bp.cell_value(i, col_ref_vmn):
            ana_list.append(bp.cell_value(i, col_ref_vmn))
        if "msr" in bp.cell_value(i, col_ref_vmn):
            msr_list.append(bp.cell_value(i, col_ref_vmn))
        if "cdp" in bp.cell_value(i, col_ref_vmn):
            cdp_list.append(bp.cell_value(i, col_ref_vmn))
        if "aku" in bp.cell_value(i, col_ref_vmn):
            aku_list.append(bp.cell_value(i, col_ref_vmn))
        if "mbm" in bp.cell_value(i, col_ref_vmn):
            mbm_list.append(bp.cell_value(i, col_ref_vmn))
        if "ttn" in bp.cell_value(i, col_ref_vmn):
            ttn_list.append(bp.cell_value(i, col_ref_vmn))
        if "mgu" in bp.cell_value(i, col_ref_vmn):
            mgu_list.append(bp.cell_value(i, col_ref_vmn))
        if "con" in bp.cell_value(i, col_ref_vmn):
            con_list.append(bp.cell_value(i, col_ref_vmn))
        if "qtp" in bp.cell_value(i, col_ref_vmn):
            qtp_list.append(bp.cell_value(i, col_ref_vmn))
        if "ccm" in bp.cell_value(i, col_ref_vmn):
            ccm_list.append(bp.cell_value(i, col_ref_vmn))
        if "qlb" in bp.cell_value(i, col_ref_vmn):
            qlb_list.append(bp.cell_value(i, col_ref_vmn))
        if "gdn" in bp.cell_value(i, col_ref_vmn):
            gdn_list.append(bp.cell_value(i, col_ref_vmn))
        if "dbm" in bp.cell_value(i, col_ref_vmn):
            dbm_list.append(bp.cell_value(i, col_ref_vmn))
        if "akf" in bp.cell_value(i, col_ref_vmn):
            akf_list.append(bp.cell_value(i, col_ref_vmn))
        if "dtl" in bp.cell_value(i, col_ref_vmn):
            dtl_list.append(bp.cell_value(i, col_ref_vmn))
        if "ssr" in bp.cell_value(i, col_ref_vmn):
            ssr_list.append(bp.cell_value(i, col_ref_vmn))
        if "vdb" in bp.cell_value(i, col_ref_vmn):
            vdb_list.append(bp.cell_value(i, col_ref_vmn))
        if "log" in bp.cell_value(i, col_ref_vmn):
            log_list.append(bp.cell_value(i, col_ref_vmn))
        if "imm" in bp.cell_value(i, col_ref_vmn):
            imm_list.append(bp.cell_value(i, col_ref_vmn))
        if "srp" in bp.cell_value(i, col_ref_vmn):
            srp_list.append(bp.cell_value(i, col_ref_vmn))
        if "crp" in bp.cell_value(i, col_ref_vmn):
            crp_list.append(bp.cell_value(i, col_ref_vmn))
        if "shd" in bp.cell_value(i, col_ref_vmn):
            shd_list.append(bp.cell_value(i, col_ref_vmn))
        if "ldr" in bp.cell_value(i, col_ref_vmn):
            ldr_list.append(bp.cell_value(i, col_ref_vmn))
        if "cgw" in bp.cell_value(i, col_ref_vmn):
            cgw_list.append(bp.cell_value(i, col_ref_vmn))
        if "dmn" in bp.cell_value(i, col_ref_vmn):
            dmn_list.append(bp.cell_value(i, col_ref_vmn))
        if "agw" in bp.cell_value(i, col_ref_vmn):
            agw_list.append(bp.cell_value(i, col_ref_vmn))
        if "cpt" in bp.cell_value(i, col_ref_vmn):
            cpt_list.append(bp.cell_value(i, col_ref_vmn))
        if "ssf" in bp.cell_value(i, col_ref_vmn):
            ssf_list.append(bp.cell_value(i, col_ref_vmn))
        if "fmc" in bp.cell_value(i, col_ref_vmn):
            fmc_list.append(bp.cell_value(i, col_ref_vmn))
        if "k8m" in bp.cell_value(i, col_ref_vmn):
            k8m_list.append(bp.cell_value(i, col_ref_vmn))
        if "k8w" in bp.cell_value(i, col_ref_vmn):
            k8w_list.append(bp.cell_value(i, col_ref_vmn))
        if "cke" in bp.cell_value(i, col_ref_vmn):
            cke_list.append(bp.cell_value(i, col_ref_vmn))
        if "keb" in bp.cell_value(i, col_ref_vmn):
            keb_list.append(bp.cell_value(i, col_ref_vmn))
        if "isb" in bp.cell_value(i, col_ref_vmn):
            isb_list.append(bp.cell_value(i, col_ref_vmn))

    # removes brackets and white spaces
    ricn_list = ('[%s]' % ','.join(map(str, ricn_list)))[1:-1]
    prb_list = ('[%s]' % ','.join(map(str, prb_list)))[1:-1]
    qrt_list = ('[%s]' % ','.join(map(str, qrt_list)))[1:-1]
    lba_list = ('[%s]' % ','.join(map(str, lba_list)))[1:-1]
    vlb_list = ('[%s]' % ','.join(map(str, vlb_list)))[1:-1]
    ana_list = ('[%s]' % ','.join(map(str, ana_list)))[1:-1]
    msr_list = ('[%s]' % ','.join(map(str, msr_list)))[1:-1]
    cdp_list = ('[%s]' % ','.join(map(str, cdp_list)))[1:-1]
    aku_list = ('[%s]' % ','.join(map(str, aku_list)))[1:-1]
    mbm_list = ('[%s]' % ','.join(map(str, mbm_list)))[1:-1]
    ttn_list = ('[%s]' % ','.join(map(str, ttn_list)))[1:-1]
    mgu_list = ('[%s]' % ','.join(map(str, mgu_list)))[1:-1]
    con_list = ('[%s]' % ','.join(map(str, con_list)))[1:-1]
    qtp_list = ('[%s]' % ','.join(map(str, qtp_list)))[1:-1]
    ccm_list = ('[%s]' % ','.join(map(str, ccm_list)))[1:-1]
    qlb_list = ('[%s]' % ','.join(map(str, qlb_list)))[1:-1]
    gdn_list = ('[%s]' % ','.join(map(str, gdn_list)))[1:-1]
    dbm_list = ('[%s]' % ','.join(map(str, dbm_list)))[1:-1]
    akf_list = ('[%s]' % ','.join(map(str, akf_list)))[1:-1]
    dtl_list = ('[%s]' % ','.join(map(str, dtl_list)))[1:-1]
    ssr_list = ('[%s]' % ','.join(map(str, ssr_list)))[1:-1]
    vdb_list = ('[%s]' % ','.join(map(str, vdb_list)))[1:-1]
    log_list = ('[%s]' % ','.join(map(str, log_list)))[1:-1]
    imm_list = ('[%s]' % ','.join(map(str, imm_list)))[1:-1]
    srp_list = ('[%s]' % ','.join(map(str, srp_list)))[1:-1]
    crp_list = ('[%s]' % ','.join(map(str, crp_list)))[1:-1]
    shd_list = ('[%s]' % ','.join(map(str, shd_list)))[1:-1]
    ldr_list = ('[%s]' % ','.join(map(str, ldr_list)))[1:-1]
    cgw_list = ('[%s]' % ','.join(map(str, cgw_list)))[1:-1]
    dmn_list = ('[%s]' % ','.join(map(str, dmn_list)))[1:-1]
    agw_list = ('[%s]' % ','.join(map(str, agw_list)))[1:-1]
    cpt_list = ('[%s]' % ','.join(map(str, cpt_list)))[1:-1]
    ssf_list = ('[%s]' % ','.join(map(str, ssf_list)))[1:-1]
    isb_list = ('[%s]' % ','.join(map(str, isb_list)))[1:-1]
    fmc_list = ('[%s]' % ','.join(map(str, fmc_list)))[1:-1]
    cke_list = ('[%s]' % ','.join(map(str, cke_list)))[1:-1]
    keb_list = ('[%s]' % ','.join(map(str, keb_list)))[1:-1]
    k8m_list = ('[%s]' % ','.join(map(str, k8m_list)))[1:-1]
    k8w_list = ('[%s]' % ','.join(map(str, k8w_list)))[1:-1]
    k8s_combined_list = k8m_list + k8w_list
    # k8s_combined_list = ('[%s]' % ','.join(map(str, k8s_combined_list)))[1:-1]

    ppp_dict = {
        "ricn_list": ricn_list,
        "prb_list": prb_list,
        "qrt_list": qrt_list,
        "lba_list": lba_list,
        "vlb_list": vlb_list,
        "ana_list": ana_list,
        "msr_list": msr_list,
        "cdp_list": cdp_list,
        "aku_list": aku_list,
        "mbm_list": mbm_list,
        "ttn_list": ttn_list,
        "mgu_list": mgu_list,
        "con_list": con_list,
        "qtp_list": qtp_list,
        "ccm_list": ccm_list,
        "qlb_list": qlb_list,
        "gdn_list": gdn_list,
        "dbm_list": dbm_list,
        "akf_list": akf_list,
        "dtl_list": dtl_list,
        "ssr_list": ssr_list,
        "vdb_list": vdb_list,
        "log_list": log_list,
        "fmc_list": fmc_list,
        "imm_list": imm_list,
        "srp_list": srp_list,
        "crp_list": crp_list,
        "shd_list": shd_list,
        "ldr_list": ldr_list,
        "cgw_list": cgw_list,
        "dmn_list": dmn_list,
        "agw_list": agw_list,
        "cpt_list": cpt_list,
        "ssf_list": ssf_list,
        "cke_list": cke_list,
        "keb_list": keb_list,
        "isb_list": isb_list,
        "k8m_list": k8m_list,
        "k8w_list": k8w_list,
        "k8s_combined_list": k8s_combined_list,
    }

    # print("ppp_dict: ", ppp_dict)

    # creates a dict of the vm-type values and the above lists

    # take vm type
    pt = openpyxl.load_workbook(preload_path)
    ws = pt[u'VMs']
    vm_type = ws['B7'].value
    vm_type = str(vm_type)
    # print("vmt: ", vm_type)

    names = []
    for k, v in ppp_dict.items():
        if k == "isb_list":
            names.append(v)

    ws = pt[u'VMs']
    vm_name = ws['C7'].value
    vm_name = vm_name[-6:]
    ppp = vm_name[0:3]
    ppp = ppp + "_list"
    ppp = str(ppp)
    # print(ppp)

    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(preload_path)
    sheet_names = wb.sheet_names()
    tag_sheet = wb.sheet_by_name(u'Tag-values')

    # search for vm type + "_names" and replace with the proper list from above
    wb = openpyxl.load_workbook(preload_path)
    ws = wb[u'Tag-values']
    for i in range(tag_sheet.nrows):
        if (tag_sheet.cell_value(i, 1) == (str(vm_type) + "_names")):
            for k, v in ppp_dict.items():
                if k == ppp:
                    val = str(i + 1)
                    # print("str(v)! :", v)
                    ws['C' + val].value = str(v)  # should this be string
    wb.save(preload_path)


def tag_sheet_indexes(preload_path, build_plan_path, count):
    """
	This function:
		- calculates the index values for the Tag-values sheet
	"""

    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(preload_path)
    sheet_names = wb.sheet_names()
    tag_sheet = wb.sheet_by_name(u'Tag-values')

    # instantiate changes to template by iterating through all rows and replacing keys with corresponding values
    wb = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = wb[u'Tag-values']
    for i in range(tag_sheet.nrows):
        if (tag_sheet.cell_value(i, 1).endswith("index")):
            val = str(i + 1)
            ws['C' + val].value = str(count)
    wb.save(preload_path)


def change_ips(preload_path, build_plan_path):
    """
	This function:
		- initiates changes for all ip related cells in Tag-Values sheet
	"""
    # open workbook and specify which sheet you would like to access
    # save vm_type
    # open workbook and specify which sheet you would like to access

    wb = xlrd.open_workbook(build_plan_path)
    sheet_names = wb.sheet_names()
    bp = wb.sheet_by_name(u'VM-Layout')

    # string search VM-Layout column headers and assign each columns reference position (int) to a variable
    # this avoids hard coding the position of certain columns
    # order does not matter

    for i in range(bp.nrows):
        for j in range(bp.ncols):
            if bp.cell_value(i, j) == "vm-name":
                col_ref_vmn = j

            if bp.cell_value(i, j) == "vm-type":
                col_ref_vmt = j

            if bp.cell_value(i, j) == "pktinternal_0_ip":
                col_ref_pk0_ip = j

            if bp.cell_value(i, j) == "pktinternal_1_ip":
                col_ref_pk1_ip = j

            if bp.cell_value(i, j) == "cdr_direct_bond_ip":
                col_ref_cdrdb_ip = j

            if bp.cell_value(i, j) == "oam_protected":
                col_ref_oam_ip = j

            if bp.cell_value(i, j) == "pktinternal_1_ip_0":
                col_ref_pktinternal_1_ip_0 = j

            if bp.cell_value(i, j) == "pktinternal_2_ip_0":
                col_ref_pktinternal_2_ip_0 = j

            if bp.cell_value(i, j) == "sd_vprobe1_ip_0":
                col_ref_sd_vprobe1_0_ip = j

            if bp.cell_value(i, j) == "sd_vprobe2_ip_0":
                col_ref_sd_vprobe2_0_ip = j

            if bp.cell_value(i, j) == "vprobe1_ip_0":
                col_ref_vprobe1_0_ip = j

            if bp.cell_value(i, j) == "vprobe2_ip_0":
                col_ref_vprobe2_0_ip = j

            if bp.cell_value(i, j) == "sd_vprobe1_cidr":
                col_ref_sd_vprobe1_cidr = j

            if bp.cell_value(i, j) == "sd_vprobe2_cidr":
                col_ref_sd_vprobe2_cidr = j

            if bp.cell_value(i, j) == "vprobe1_cidr":
                col_ref_vprobe1_cidr = j

            if bp.cell_value(i, j) == "vprobe2_cidr":
                col_ref_vprobe2_cidr = j

            if bp.cell_value(i, j) == "cdr_direct_bond_ip":
                col_ref_cdbi = j

            if bp.cell_value(i, j) == "vertica_ic_bond_ip":
                col_ref_vibi = j

            if bp.cell_value(i, j) == "VLAN_ID_LEFT_ip_0":
                col_ref_vili0 = j

            if bp.cell_value(i, j) == "VLAN_ID_LEFT_mac":
                col_ref_vilm = j

            if bp.cell_value(i, j) == "VLAN_ID_RIGHT_ip_0":
                col_ref_viri0 = j

            if bp.cell_value(i, j) == "VLAN_ID_RIGHT_mac":
                col_ref_virm = j

            if bp.cell_value(i, j) == "ngxp_oam_net":
                col_ref_oam_net_ip = j

            if bp.cell_value(i, j) == "ngxp_mgt_net":
                col_ref_mgt_ip = j

            if bp.cell_value(i, j) == "ngxp_cdr_net":
                col_ref_cdr_ip = j

            if bp.cell_value(i, j) == "int_ngxp_be_ccd_net":
                col_ref_be_ccd_ip = j

            if bp.cell_value(i, j) == "int_ngxp_admin_ccd_net":
                col_ref_admin_ccd_ip = j

            if bp.cell_value(i, j) == "int_ngxp_ecde_ccd_net":
                col_ref_ecde_ccd_ip = j

            if bp.cell_value(i, j) == "oam_protected_net":
                col_ref_oam_protected_ip = j

            if bp.cell_value(i, j) == "int_ngxp_fe_ccd_net":
                col_ref_fe_ccd_ip = j

            if bp.cell_value(i, j) == "int_ngxp_pcu_net":
                col_ref_fe_pcu_ip = j

            if bp.cell_value(i, j) == "int_probe_agent_net_1":
                col_ref_fe_net1_ip = j

            if bp.cell_value(i, j) == "int_probe_agent_net_2":
                col_ref_fe_net2_ip = j

    # create dictionaries of vm-names and ip's

    vili0_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            vili0 = bp.cell_value(i, col_ref_vili0)
            if vm_name in vili0_dict:
                vili0_dict[vm_name].append(vili0)
            else:
                vili0_dict[vm_name] = [vili0]
        except:
            pass
    vili0_dict = {k: ",".join(str(x) for x in v) for k, v in vili0_dict.items()}

    vilm_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            vilm = bp.cell_value(i, col_ref_vilm)
            if vm_name in vilm_dict:
                vilm_dict[vm_name].append(vilm)
            else:
                vilm_dict[vm_name] = [vilm]
        except:
            pass
    vilm_dict = {k: ",".join(str(x) for x in v) for k, v in vilm_dict.items()}

    viri0_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            viri0 = bp.cell_value(i, col_ref_viri0)
            if vm_name in viri0_dict:
                viri0_dict[vm_name].append(viri0)
            else:
                viri0_dict[vm_name] = [viri0]
        except:
            pass
    viri0_dict = {k: ",".join(str(x) for x in v) for k, v in viri0_dict.items()}

    virm_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            virm = bp.cell_value(i, col_ref_virm)
            if vm_name in virm_dict:
                virm_dict[vm_name].append(virm)
            else:
                virm_dict[vm_name] = [virm]
        except:
            pass
    virm_dict = {k: ",".join(str(x) for x in v) for k, v in virm_dict.items()}

    vprobe1_cidr_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            vp1cidr = bp.cell_value(i, col_ref_vprobe1_cidr)
            if vm_name in vprobe1_cidr_dict:
                vprobe1_cidr_dict[vm_name].append(vp1cidr)
            else:
                vprobe1_cidr_dict[vm_name] = [vp1cidr]
        except:
            pass
    vprobe1_cidr_dict = {k: ",".join(str(x) for x in v) for k, v in vprobe1_cidr_dict.items()}

    vprobe2_cidr_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            vp2cidr = bp.cell_value(i, col_ref_vprobe2_cidr)
            if vm_name in vprobe2_cidr_dict:
                vprobe1_cidr_dict[vm_name].append(vp2cidr)
            else:
                vprobe1_cidr_dict[vm_name] = [vp2cidr]
        except:
            pass
    vprobe2_cidr_dict = {k: ",".join(str(x) for x in v) for k, v in vprobe2_cidr_dict.items()}

    sd_vprobe1_cidr_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            vp1cidr = bp.cell_value(i, col_ref_sd_vprobe1_cidr)
            if vm_name in sd_vprobe1_cidr_dict:
                vprobe1_cidr_dict[vm_name].append(vp1cidr)
            else:
                vprobe1_cidr_dict[vm_name] = [vp1cidr]
        except:
            pass
    sd_vprobe1_cidr_dict = {k: ",".join(str(x) for x in v) for k, v in sd_vprobe1_cidr_dict.items()}

    sd_vprobe2_cidr_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            vp2cidr = bp.cell_value(i, col_ref_sd_vprobe2_cidr)
            if vm_name in sd_vprobe2_cidr_dict:
                vprobe1_cidr_dict[vm_name].append(vp2cidr)
            else:
                vprobe1_cidr_dict[vm_name] = [vp2cidr]
        except:
            pass
    sd_vprobe2_cidr_dict = {k: ",".join(str(x) for x in v) for k, v in sd_vprobe2_cidr_dict.items()}

    pktinternal_1_ip_0_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            pkt = bp.cell_value(i, col_ref_pktinternal_1_ip_0)
            if vm_name in pktinternal_1_ip_0_dict:
                pktinternal_1_ip_0_dict[vm_name].append(pkt)
            else:
                pktinternal_1_ip_0_dict[vm_name] = [pkt]
        except:
            pass
    pktinternal_1_ip_0_dict = {k: ",".join(str(x) for x in v) for k, v in pktinternal_1_ip_0_dict.items()}

    pktinternal_2_ip_0_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            pkt = bp.cell_value(i, col_ref_pktinternal_2_ip_0)
            if vm_name in pktinternal_2_ip_0_dict:
                pktinternal_2_ip_0_dict[vm_name].append(pkt)
            else:
                pktinternal_2_ip_0_dict[vm_name] = [pkt]
        except:
            pass
    pktinternal_2_ip_0_dict = {k: ",".join(str(x) for x in v) for k, v in pktinternal_2_ip_0_dict.items()}

    pkt0_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            pk0_ip = bp.cell_value(i, col_ref_pk0_ip)
            if vm_name in pkt0_dict:
                pkt0_dict[vm_name].append(pk0_ip)
            else:
                pkt0_dict[vm_name] = [pk0_ip]
        except:
            pass
    pkt0_dict = {k: ",".join(str(x) for x in v) for k, v in pkt0_dict.items()}

    pkt1_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            pk1_ip = bp.cell_value(i, col_ref_pk1_ip)
            if vm_name in pkt1_dict:
                pkt1_dict[vm_name].append(pk1_ip)
            else:
                pkt1_dict[vm_name] = [pk1_ip]
        except:
            pass
    pkt1_dict = {k: ",".join(str(x) for x in v) for k, v in pkt1_dict.items()}

    cdr_direct_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            cdr = bp.cell_value(i, col_ref_cdrdb_ip)
            if vm_name in cdr_direct_dict:
                cdr_direct_dict[vm_name].append(cdr)
            else:
                cdr_direct_dict[vm_name] = [cdr]
        except:
            pass
    cdr_direct_dict = {k: ",".join(str(x) for x in v) for k, v in cdr_direct_dict.items()}

    oam_dict = {}
    for i in range(bp.nrows):
        try:
            vm_type = bp.cell_value(i, col_ref_vmt)
            oam_ips = bp.cell_value(i, col_ref_oam_ip)
            if vm_type in oam_dict:
                oam_dict[vm_type].append(oam_ips)
            else:
                oam_dict[vm_type] = [oam_ips]
        except:
            pass
    oam_dict = {k: ",".join(str(x) for x in v) for k, v in oam_dict.items()}

    sd_vprobe1_0_ip_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            sd_vprobe1_0_ips = bp.cell_value(i, col_ref_sd_vprobe1_0_ip)
            if vm_name in sd_vprobe1_0_ip_dict:
                sd_vprobe1_0_ip_dict[vm_name].append(sd_vprobe1_0_ips)
            else:
                sd_vprobe1_0_ip_dict[vm_name] = [sd_vprobe1_0_ips]
        except:
            pass
    sd_vprobe1_0_ip_dict = {k: ",".join(str(x) for x in v) for k, v in sd_vprobe1_0_ip_dict.items()}

    sd_vprobe2_0_ip_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            sd_vprobe2_0_ips = bp.cell_value(i, col_ref_sd_vprobe2_0_ip)
            if vm_name in sd_vprobe2_0_ip_dict:
                sd_vprobe2_0_ip_dict[vm_name].append(sd_vprobe2_0_ips)
            else:
                sd_vprobe2_0_ip_dict[vm_name] = [sd_vprobe2_0_ips]
        except:
            pass
    sd_vprobe2_0_ip_dict = {k: ",".join(str(x) for x in v) for k, v in sd_vprobe2_0_ip_dict.items()}

    vprobe1_0_ip_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            # print("VM NAME: ", vm_name)
            vprobe1_ip_0_ips = bp.cell_value(i, col_ref_vprobe1_0_ip)
            # print("IPS : ", vprobe1_ip_0_ips)
            if vm_name in vprobe1_0_ip_dict:
                vprobe1_0_ip_dict[vm_name].append(vprobe1_ip_0_ips)
            else:
                vprobe1_0_ip_dict[vm_name] = [vprobe1_ip_0_ips]
        except:
            pass
    vprobe1_0_ip_dict = {k: ",".join(str(x) for x in v) for k, v in vprobe1_0_ip_dict.items()}

    vprobe2_0_ip_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            vprobe2_ip_0_ips = bp.cell_value(i, col_ref_vprobe2_0_ip)
            if vm_name in vprobe2_0_ip_dict:
                vprobe2_0_ip_dict[vm_name].append(vprobe2_ip_0_ips)
            else:
                vprobe2_0_ip_dict[vm_name] = [vprobe2_ip_0_ips]
        except:
            pass
    vprobe2_0_ip_dict = {k: ",".join(str(x) for x in v) for k, v in vprobe2_0_ip_dict.items()}

    cdr_direct_bond_ip_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            cdbi_ips = bp.cell_value(i, col_ref_cdbi)
            if vm_name in cdr_direct_bond_ip_dict:
                cdr_direct_bond_ip_dict[vm_name].append(cdbi_ips)
            else:
                cdr_direct_bond_ip_dict[vm_name] = [cdbi_ips]
        except:
            pass
    cdr_direct_bond_ip_dict = {k: ",".join(str(x) for x in v) for k, v in cdr_direct_bond_ip_dict.items()}

    vertica_ic_bond_ip_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            vibi_ips = bp.cell_value(i, col_ref_vibi)
            if vm_name in vertica_ic_bond_ip_dict:
                vertica_ic_bond_ip_dict[vm_name].append(vibi_ips)
            else:
                vertica_ic_bond_ip_dict[vm_name] = [vibi_ips]
        except:
            pass
    vertica_ic_bond_ip_dict = {k: ",".join(str(x) for x in v) for k, v in vertica_ic_bond_ip_dict.items()}

    ngxp_oam_net_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_oam = bp.cell_value(i, col_ref_oam_net_ip)
            if vm_name in ngxp_oam_net_dict:
                ngxp_oam_net_dict[vm_name].append(ngxp_oam)
            else:
                ngxp_oam_net_dict[vm_name] = [ngxp_oam]
        except:
            pass
    ngxp_oam_net_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_oam_net_dict.items()}

    ngxp_mgt_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_mgt = bp.cell_value(i, col_ref_mgt_ip)
            if vm_name in ngxp_mgt_dict:
                ngxp_mgt_dict[vm_name].append(ngxp_mgt)
            else:
                ngxp_mgt_dict[vm_name] = [ngxp_mgt]
        except:
            pass
    ngxp_mgt_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_mgt_dict.items()}

    ngxp_cdr_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_cdr = bp.cell_value(i, col_ref_cdr_ip)
            if vm_name in ngxp_cdr_dict:
                ngxp_cdr_dict[vm_name].append(ngxp_cdr)
            else:
                ngxp_cdr_dict[vm_name] = [ngxp_cdr]
        except:
            pass
    ngxp_cdr_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_cdr_dict.items()}

    ngxp_be_ccd_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_be_ccd = bp.cell_value(i, col_ref_be_ccd_ip)
            if vm_name in ngxp_be_ccd_dict:
                ngxp_be_ccd_dict[vm_name].append(ngxp_be_ccd)
            else:
                ngxp_be_ccd_dict[vm_name] = [ngxp_be_ccd]
        except:
            pass
    ngxp_be_ccd_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_be_ccd_dict.items()}

    ngxp_admin_ccd_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_admin_ccd = bp.cell_value(i, col_ref_admin_ccd_ip)
            if vm_name in ngxp_admin_ccd_dict:
                ngxp_admin_ccd_dict[vm_name].append(ngxp_admin_ccd)
            else:
                ngxp_admin_ccd_dict[vm_name] = [ngxp_admin_ccd]
        except:
            pass
    ngxp_admin_ccd_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_admin_ccd_dict.items()}

    ngxp_ecde_ccd_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_ecde_ccd = bp.cell_value(i, col_ref_ecde_ccd_ip)
            if vm_name in ngxp_ecde_ccd_dict:
                ngxp_admin_ccd_dict[vm_name].append(ngxp_ecde_ccd)
            else:
                ngxp_ecde_ccd_dict[vm_name] = [ngxp_ecde_ccd]
        except:
            pass
    ngxp_ecde_ccd_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_ecde_ccd_dict.items()}

    ngxp_oam_protected_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_oam_protected = bp.cell_value(i, col_ref_oam_protected_ip)
            if vm_name in ngxp_oam_protected_dict:
                ngxp_oam_protected_dict[vm_name].append(ngxp_oam_protected)
            else:
                ngxp_oam_protected_dict[vm_name] = [ngxp_oam_protected]
        except:
            pass
    ngxp_oam_protected_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_oam_protected_dict.items()}

    ngxp_fe_ccd_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_fe_ccd = bp.cell_value(i, col_ref_fe_ccd_ip)
            if vm_name in ngxp_fe_ccd_dict:
                ngxp_fe_ccd_dict[vm_name].append(ngxp_fe_ccd)
            else:
                ngxp_fe_ccd_dict[vm_name] = [ngxp_fe_ccd]
        except:
            pass
    ngxp_fe_ccd_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_fe_ccd_dict.items()}

    ngxp_fe_pcu_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_fe_pcu = bp.cell_value(i, col_ref_fe_pcu_ip)
            if vm_name in ngxp_fe_pcu_dict:
                ngxp_fe_pcu_dict[vm_name].append(ngxp_fe_pcu)
            else:
                ngxp_fe_pcu_dict[vm_name] = [ngxp_fe_pcu]
        except:
            pass
    ngxp_fe_pcu_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_fe_pcu_dict.items()}

    ngxp_fe_net1_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_fe_net1 = bp.cell_value(i, col_ref_fe_net1_ip)
            if vm_name in ngxp_fe_net1_dict:
                ngxp_fe_net1_dict[vm_name].append(ngxp_fe_net1)
            else:
                ngxp_fe_net1_dict[vm_name] = [ngxp_fe_net1]
        except:
            pass
    ngxp_fe_net1_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_fe_net1_dict.items()}

    ngxp_fe_net2_dict = {}
    for i in range(bp.nrows):
        try:
            vm_name = bp.cell_value(i, col_ref_vmn)
            ngxp_fe_net2 = bp.cell_value(i, col_ref_fe_net2_ip)
            if vm_name in ngxp_fe_net2_dict:
                ngxp_fe_net2_dict[vm_name].append(ngxp_fe_net2)
            else:
                ngxp_fe_net2_dict[vm_name] = [ngxp_fe_net2]
        except:
            pass
    ngxp_fe_net2_dict = {k: ",".join(str(x) for x in v) for k, v in ngxp_fe_net2_dict.items()}

    # create dictionary of dictionaries
    ip_dict = {
        "cdr_direct_bond_ip": cdr_direct_dict,
        "oam_protected_ips": oam_dict,
        "pktinternal_1_ip_0": pktinternal_1_ip_0_dict,
        "pktinternal_2_ip_0": pktinternal_2_ip_0_dict,
        "sd_vprobe1_ip_0": sd_vprobe1_0_ip_dict,
        "vprobe1_ip_0": vprobe1_0_ip_dict,
        "vprobe2_ip_0": vprobe2_0_ip_dict,
        "sd_vprobe2_ip_0": sd_vprobe2_0_ip_dict,
        "cdr_direct_bond_ip": cdr_direct_bond_ip_dict,
        "vertica_ic_bond_ip": vertica_ic_bond_ip_dict,
        "VLAN_ID_LEFT_ip_0": vili0_dict,
        "VLAN_ID_LEFT_mac": vilm_dict,
        "VLAN_ID_RIGHT_ip_0": viri0_dict,
        "VLAN_ID_RIGHT_mac": virm_dict,
        "ngxp_oam1_ip_0": ngxp_oam_net_dict,
        "ngxp_mgt1_ip_0": ngxp_mgt_dict,
        "ngxp_cdr1_ip_0": ngxp_cdr_dict,
        "ngxp_be_ccd1_ip_0": ngxp_be_ccd_dict,
        "ngxp_admin_ccd1_ip_0": ngxp_admin_ccd_dict,
        "ngxp_ecde_ccd1_ip_0": ngxp_ecde_ccd_dict,
        "oam_protected1_ip_0": ngxp_oam_protected_dict,
        "ngxp_fe_ccd1_ip_0": ngxp_fe_ccd_dict,
        "ngxp_fe_pcu_ip_0": ngxp_fe_pcu_dict,
        "ngxp_fe_net_1_ip_0": ngxp_fe_net1_dict,
        "ngxp_fe_net_2_ip_0": ngxp_fe_net2_dict

    }

    kuberiq_ips_combined = []
    for k, v in ip_dict["oam_protected_ips"].items():
        if k and "kuberiq" in k:  # skip float or empty string
            kuberiq_ips_combined.append(v)
    if len(kuberiq_ips_combined) > 0:
        k_stripped = ('[%s]' % ','.join(map(str, kuberiq_ips_combined)))[1:-1]

    # open workbook and specify which sheet you would like to access
    wb = xlrd.open_workbook(preload_path)
    sheet_names = wb.sheet_names()
    tag_sheet = wb.sheet_by_name(u'Tag-values')

    pt = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = pt[u'VMs']
    vm_type = ws['B7'].value
    vm_type = str(vm_type)

    vm_name = ws['C7'].value
    vm_name = str(vm_name)

    wb = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = wb[u'Tag-values']
    for i in range(tag_sheet.nrows):
        for k, v in ip_dict.items():
            if re.findall(k, tag_sheet.cell_value(i, 1)):
                for k1, v1 in v.items():
                    if k1 == vm_type or k1 == vm_name:
                        val = str(i + 1)
                        ws['C' + val].value = v1
            if tag_sheet.cell_value(i, 1).startswith("sd_vprobe1_ip_0"):
                for k1, v1 in sd_vprobe1_0_ip_dict.items():
                    if k1 == vm_name:
                        val = str(i + 1)
                        ws['C' + val].value = v1
            if tag_sheet.cell_value(i, 1).startswith("sd_vprobe2_ip_0"):
                for k1, v1 in sd_vprobe2_0_ip_dict.items():
                    if k1 == vm_name:
                        val = str(i + 1)
                        ws['C' + val].value = v1
            if tag_sheet.cell_value(i, 1).startswith("kuberiq") and tag_sheet.cell_value(i, 1).endswith(
                    "oam_protected_ips"):
                val = str(i + 1)
                ws['C' + val].value = k_stripped

    wb.save(preload_path)


def change_ngxp_tag(preload_path, build_plan_path):
    """
	This function:
		- fills in the ngxp values for Tag-values sheet
	"""

    # open workbook and specify which sheet you would like to access
    wb = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    ws = wb[u'Tag-values']

    # only fills in Tag-values for ngxp preloads
    try:
        is_ngxp_preload = [i for i, col in enumerate(ws.iter_cols()) for cell in col if cell.value == "vnf_role"][0]

        wp = openpyxl.load_workbook(build_plan_path, data_only=True, keep_vba=True)
        bp = wp[u'VM-Layout']
        col_ref_vmn = [i for i, col in enumerate(bp.iter_cols()) for cell in col if cell.value == "vm-name"][0]
        bv = wp[u'VNF-Specs']
        col_ref_vmt = [i for i, col in enumerate(bv.iter_cols()) for cell in col if cell.value == "vm-type"][0]
        col_ref_flavor = [i for i, col in enumerate(bv.iter_cols()) for cell in col if str(cell.value).startswith("Flavor")][0]
        col_ref_image = [i for i, col in enumerate(bv.iter_cols()) for cell in col if str(cell.value).startswith("Image")][0]

        vm_name = str(wb[u'VMs']['C7'].value)
        vm_type = str(wb[u'VMs']['B7'].value)
        ppp = vm_name[-6:-3]
        col_ref_tag = [i for i, col in enumerate(ws.iter_cols()) for cell in col if cell.value == "vnf-parameter-name"][0]
        flavor_name = [row[col_ref_flavor].value for row in bv.rows if row[col_ref_vmt].value == vm_type][0]
        image_name = [row[col_ref_image].value for row in bv.rows if row[col_ref_vmt].value == vm_type][0]

        # create dict of vm-type cluster indexing values of VM Count
        be_ccd_list = ['bcd','bcm','bcw','bcc','bcu']
        admin_ccd_list = ['acd','acm','acw']
        ecde_ccd_list = ['ecd','ecm','ecw']
        be_ccd_dict, admin_ccd_dict, ecde_ccd_dict = ({} for i in range(3))
        idx_bcd, idx_adm, idx_ecd = (0 for i in range(3))
        for row in bp.rows:
            if row[col_ref_vmn].value and row[col_ref_vmn].value != "vm-name":
                ppp_val = row[col_ref_vmn].value[-6:-3]
                if ppp_val in be_ccd_list:
                    idx_bcd += 1
                    be_ccd_dict[row[col_ref_vmn].value] = idx_bcd
                elif ppp_val in admin_ccd_list:
                    idx_adm += 1
                    admin_ccd_dict[row[col_ref_vmn].value] = idx_adm
                elif ppp_val in ecde_ccd_list:
                    idx_ecd += 1
                    ecde_ccd_dict[row[col_ref_vmn].value] = idx_ecd

        # fill in the Tag-values
        for i, row in enumerate(ws.rows):
            val = str(row[col_ref_tag].value)
            if val.endswith("vnf_role"):
                ws.cell(i + 1, col_ref_tag + 2).value = ppp
            if val == "vnf_index":
                if ppp in be_ccd_list:
                    ws.cell(i + 1, col_ref_tag + 2).value = [idx for key,idx in be_ccd_dict.items() if key == vm_name][0]
                elif ppp in admin_ccd_list:
                    ws.cell(i + 1, col_ref_tag + 2).value = [idx for key,idx in admin_ccd_dict.items() if key == vm_name][0]
                elif ppp in ecde_ccd_list:
                    ws.cell(i + 1, col_ref_tag + 2).value = [idx for key,idx in ecde_ccd_dict.items() if key == vm_name][0]
                else:
                    ws.cell(i + 1, col_ref_tag + 2).value = int(vm_name[-3:])
            if val.endswith("workload_context"):
                ws.cell(i + 1, col_ref_tag + 2).value = vm_name + "_work_context"
            if val.endswith("environment_context"):
                ws.cell(i + 1, col_ref_tag + 2).value = vm_name + "_env_context"
            if val.endswith("name_0"):
                ws.cell(i + 1, col_ref_tag + 2).value = vm_name
            if val.endswith("flavor_name"):
                ws.cell(i + 1, col_ref_tag + 2).value = flavor_name
            if val.endswith("image_name"):
                ws.cell(i + 1, col_ref_tag + 2).value = image_name

        wb.save(preload_path)
    except IndexError:
        pass


def delete_image(preload_path):
    wb1 = openpyxl.load_workbook(preload_path)
    wb1.save(preload_path)
    pt = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
    #wb = pt[u'Instructions']
    #wb.delete_rows(19, 20)
    #wb._images = []  # delete all images
    pt.save(preload_path)


start = datetime.datetime.now()

build_plan_path = sys.argv[1]
preload_list = []
paths = sys.argv[2]
for idx, item in enumerate(os.listdir(paths)):
    preload_list.append(paths + item)

dest_folder = sys.argv[3]

files = []
for preload_path in preload_list:
    delete_image(preload_path)
    if "base" not in preload_path:
        calculate_vm_count(build_plan_path)
        for titles in range(0, len(title_list)):
            change_general(preload_path, build_plan_path, str(titles + 1))
            change_networks(preload_path, build_plan_path)
            change_probe_prod(preload_path, build_plan_path)
            change_vccn(preload_path, build_plan_path)
            change_tag(preload_path, build_plan_path)
            change_vm(preload_path, build_plan_path, str(titles + 1))
            change_az(preload_path, build_plan_path)
            change_vm_network_ips(preload_path, build_plan_path)
            names_tag_sheet_nf_naming(preload_path, build_plan_path)
            # names_tag_sheet(preload_path, build_plan_path)
            tag_sheet_indexes(preload_path, build_plan_path, str(titles))
            change_ips(preload_path, build_plan_path)
            change_ngxp_tag(preload_path, build_plan_path)

            wb = openpyxl.load_workbook(preload_path, read_only=False, keep_vba=True)
            wb.save(dest_folder + title_list[titles] + str(titles + 1) + ".xlsm")

# datetime object containing current date and time

now = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
# name = dest_folder.rsplit('/')[-2]
name = os.path.basename(os.path.dirname(dest_folder))
zipf = name  # + "-" + str(now)
archive_name = os.path.expanduser(os.path.join(dest_folder + zipf))
root_dir = os.path.expanduser(os.path.join(dest_folder))
make_archive(archive_name, 'tar', root_dir)

for fname in os.listdir(dest_folder):
    if fname.endswith(".xlsm"):
        os.remove(os.path.join(dest_folder, fname))

#print(f"\n--- {datetime.datetime.now() - start} seconds ---")
